from openpyxl import load_workbook

file = input("Enter the file path: ") + ".xlsx"
try:
    workbook = load_workbook(file)
    sheet = workbook.active
    publisher = input("Enter the publisher name: ")
    # book_list = []
    # publisher_list = []
    for i in range(2, sheet.max_row + 1):
        n = f"D{i}"
        b = f"B{i}"
        if publisher in sheet[n].value:
            sheet[n] = publisher
            # publisher_list.append(sheet[n].value)
            # book_list.append(sheet[b].value)
        else:
            continue

    workbook.save("saved.xlsx")
    # print("List of books whos' Publishers are changed: ", book_list)
    # print(publisher_list)

except FileNotFoundError:
    print("Please enter a valid file path!")


