from openpyxl import load_workbook
from tkinter import filedialog

file = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Any file", "*")))

try:
    workbook = load_workbook(file)
    sheet = workbook.active
    # publisher = input("Enter the publisher name: ")
    # book_list = []
    # publisher_list = []
    ColNames = {}
    Current = 0
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
    # print(ColNames)
    alphabets = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for name in ColNames.keys():
        if "publish" in name.lower():
            col = alphabets[ColNames[name]]

    for i in range(2, sheet.max_row + 1):
        try:
            n = f"{col}{i}"
            # b = f"B{i}"

            publisher = sheet[n].value
            for j in range(3, sheet.max_row + 1):
                n1 = f"{col}{j}"
                # b1 = f"B{i}"
                if publisher in sheet[n1].value:
                    sheet[n1] = publisher
                    # publisher_list.append(sheet[n1].value)
                    # book_list.append(sheet[b1].value)
                else:
                    continue
        except NameError:
            print("'Publisher' column is not found. Make sure it is named properly in the Excel sheet")

    workbook.save("saved.xlsx")
    # print("List of books whos' Publishers are changed: ", book_list)
    # print(publisher_list)

except FileNotFoundError:
    print("Please enter a valid file path!")


